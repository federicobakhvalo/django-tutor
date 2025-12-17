from django.db.models import F, Q
from django.urls import reverse_lazy
from django.views.generic import TemplateView, CreateView,ListView,UpdateView
from openpyxl.reader.excel import load_workbook

from django.conf import settings
from .models import *
from .forms import *
from django.db import transaction
from .models_utils.mixins import TitleContextMixin

# Create your views here.


class MainPageView(TemplateView):
    template_name = "main.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['menu_items'] = [
            {'title': 'Все книги', 'url': '/books/'},
            {'title': 'Предложить книгу', 'url': '/create_book/'},
            {'title': 'Создать читателя', 'url': '/create_reader/'},
            {'title': 'Выдача книг', 'url': '/bookloan_history/'},
            {'title':"Создать читательский билет",'url':"/create_ticket/"}
        ]
        return context


class CreateBookView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'forms/form.html'
    success_url = reverse_lazy('main')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Предложить книгу'
        return context
#


class CreateReaderView(TitleContextMixin,CreateView):
    model=Reader
    form_class = ReaderForm
    template_name = 'forms/form.html'
    success_url = reverse_lazy('main')
    title = 'Создать читателя'




class CreateBookLoan(TitleContextMixin,CreateView):
    model=BookLoan
    template_name = 'forms/form.html'
    form_class = BookLoanForm
    success_url = reverse_lazy('main')
    title = 'Забронировать книгу'


    def get_initial(self):
        initial = super().get_initial()
        book_id = self.request.GET.get('book_to')
        if book_id:
            initial['book'] = book_id
        return initial





class BookListView(ListView):
    model = Book
    template_name = 'books/book_list.html'
    context_object_name = 'books'
    paginate_by = 10

    def get_queryset(self):
        queryset = (
            Book.objects
            .select_related('author')
            .annotate(author_name=F('author__name'))
        )
        self.q = self.request.GET.get('q', '').strip()
        if self.q:
            queryset = queryset.filter(Q(bookname__icontains=self.q) | Q(author_name__icontains = self.q))
        return queryset.order_by('bookname')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = self.q
        return context


class BookLoanListView(ListView):
    model = BookLoan
    template_name = 'books/book_loan.html'
    context_object_name = 'loans'
    paginate_by = 10

    def get_queryset(self):
        queryset = BookLoan.objects.select_related('book', 'reader', 'librarian')
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(book__bookname__icontains=q) |
                Q(reader__first_name__icontains=q) |
                Q(reader__last_name__icontains=q)
            )
        queryset = queryset.order_by('-issued_at')
        return queryset


class UpdateBookLoanView(TitleContextMixin,UpdateView):
    model = BookLoan
    form_class = BookLoanUpdateForm
    template_name = 'forms/form.html'
    success_url = reverse_lazy('bookloan_history')
    title = 'Обновить выдачу книги'



class CreateTicketView(TitleContextMixin,CreateView):
    model=ReaderTicket
    form_class =ReaderTicketForm
    template_name = 'forms/form.html'
    success_url = reverse_lazy('main')
    title = 'Создать читательский билет'

class BookLoanExcelView(TemplateView):
    template_name = 'books/bookloan_excel.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        file_path = settings.BASE_DIR / 'data' / 'book_loans.xlsx'
        wb = load_workbook(file_path)
        ws = wb.active
        rows = []
        book_ids, reader_ids, librarian_ids = set(), set(), set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(row)
            book_ids.add(row[0])
            reader_ids.add(row[1])
            librarian_ids.add(row[2])

        books = Book.objects.filter(id__in=book_ids).values('id', 'bookname')
        readers = Reader.objects.filter(id__in=reader_ids).values('id', 'first_name', 'last_name')
        librarians = Librarian.objects.filter(id__in=librarian_ids).values('id', 'first_name', 'last_name')
        book_map = {b['id']: b['bookname'] for b in books}
        reader_map = {r['id']: f"{r['first_name']} {r['last_name']}" for r in readers}
        librarian_map = {l['id']: f"{l['first_name']} {l['last_name']}" for l in librarians}

        loans = []
        for row in rows:
            loans.append({
                'book': book_map.get(row[0], '—'),
                'reader': reader_map.get(row[1], '—'),
                'librarian': librarian_map.get(row[2], '—'),
                'due_date': row[3],
                'returned_at': row[4],
            })
        context['loans'] = loans
        context['title'] = 'История выдач (Excel)'
        return context



